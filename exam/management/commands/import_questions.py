import re

import openpyxl
from django.core.management.base import BaseCommand
from pathlib import Path

from exam.models import Exam, Question
from main.models import Subject


EXAM_TYPE_MAP = {
    '기말시험': '기말',
    '기말': '기말',
    '중간시험': '중간',
    '중간': '중간',
    '계절학기': '계절',
    '계절': '계절',
}

# 중복답안 A~K → "1,2" 등 매핑
MULTI_ANSWER_MAP = {
    'A': '1,2', 'B': '1,3', 'C': '1,4',
    'D': '2,3', 'E': '2,4', 'F': '3,4',
    'G': '1,2,3', 'H': '1,2,4', 'I': '1,3,4',
    'J': '2,3,4', 'K': '1,2,3,4',
}

# ── 유니코드 상·하첨자 변환 ──────────────────────────────
SUB_MAP = str.maketrans('0123456789', '₀₁₂₃₄₅₆₇₈₉')
SUP_MAP = str.maketrans('0123456789+-', '⁰¹²³⁴⁵⁶⁷⁸⁹⁺⁻')


def _to_sub(s):
    return s.translate(SUB_MAP)


def _to_sup(s):
    return s.translate(SUP_MAP)


def convert_formulas(text):
    """화학식·단위의 상하첨자를 유니코드로 변환한다.

    변환 예시:
        H2O → H₂O,  CO2 → CO₂,  Ca2+ → Ca²⁺,
        PO43- → PO₄³⁻,  NO3- → NO₃⁻,  cm3 → cm³
    """
    if not text:
        return ''
    text = str(text)

    # ① 단위 지수: cm3 → cm³, mm2 → mm², km2 → km² (공백 허용)
    text = re.sub(
        r'(cm|mm|km)\s*(\d)',
        lambda m: m.group(1) + _to_sup(m.group(2)),
        text,
    )
    # 독립 m + 숫자: m2 → m², m3 → m³, m 2 → m² (공백 허용)
    text = re.sub(
        r'(?<![a-zA-Z])m\s*(\d)(?!\d)',
        lambda m: 'm' + _to_sup(m.group(1)),
        text,
    )

    # ② 다가 이온: SO42- → SO₄²⁻, PO43- → PO₄³⁻ (공백 허용)
    text = re.sub(
        r'([A-Z][a-z]?)\s*(\d)\s*(\d)\s*([+-])',
        lambda m: m.group(1) + _to_sub(m.group(2)) + _to_sup(m.group(3) + m.group(4)),
        text,
    )

    # ③ 2글자 원소 이온: Ca2+ → Ca²⁺, Fe 2+ → Fe²⁺ (공백 허용)
    text = re.sub(
        r'([A-Z][a-z])\s*(\d)\s*([+-])',
        lambda m: m.group(1) + _to_sup(m.group(2) + m.group(3)),
        text,
    )

    # ④ 원소 뒤 부호만: K+ → K⁺, H+ → H⁺, Cl- → Cl⁻
    text = re.sub(
        r'([A-Z][a-z]?)([+-])(?=[\s,;.)\]}]|$)',
        lambda m: m.group(1) + _to_sup(m.group(2)),
        text,
    )

    # ⑤ 화학식 아래첨자: H2O → H₂O, NH 2 → NH₂ (공백 허용)
    text = re.sub(
        r'([A-Z][a-z]?)\s*(\d+)',
        lambda m: m.group(1) + _to_sub(m.group(2)),
        text,
    )

    # ⑥ 아래첨자 뒤 남은 전하 부호: NO₃- → NO₃⁻ (공백 허용)
    text = re.sub(
        r'([\u2080-\u2089])\s*([+-])',
        lambda m: m.group(1) + _to_sup(m.group(2)),
        text,
    )

    # ⑦ 괄호 뒤 아래첨자: (OH)2 → (OH)₂ (공백 허용)
    text = re.sub(
        r'\)\s*(\d+)',
        lambda m: ')' + _to_sub(m.group(1)),
        text,
    )

    return text


class Command(BaseCommand):
    help = 'data 디렉토리의 엑셀 파일에서 기출문제를 import합니다.'

    def add_arguments(self, parser):
        parser.add_argument('file', nargs='?', help='특정 파일명 (예: 토양학.xlsx). 생략하면 data/ 전체.')

    def handle(self, *args, **options):
        data_dir = Path('data')
        if not data_dir.exists():
            self.stderr.write(self.style.ERROR('data 디렉토리가 없습니다.'))
            return

        if options['file']:
            files = [data_dir / options['file']]
        else:
            files = sorted(data_dir.glob('*.xlsx'))

        if not files:
            self.stderr.write(self.style.ERROR('엑셀 파일이 없습니다.'))
            return

        total_exams = 0
        total_questions = 0
        total_updated = 0

        for filepath in files:
            if not filepath.exists():
                self.stderr.write(self.style.WARNING(f'파일 없음: {filepath}'))
                continue

            self.stdout.write(f'\n처리 중: {filepath.name}')
            exams_created, questions_created, questions_updated = self._import_file(filepath)
            total_exams += exams_created
            total_questions += questions_created
            total_updated += questions_updated
            self.stdout.write(self.style.SUCCESS(
                f'  → 시험 {exams_created}개, 문제 {questions_created}개 생성, {questions_updated}개 갱신'
            ))

        self.stdout.write(self.style.SUCCESS(
            f'\n완료: 총 시험 {total_exams}개, 문제 {total_questions}개 생성, {total_updated}개 갱신'
        ))

    def _import_file(self, filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        col = {h: i for i, h in enumerate(headers)}

        exams_created = 0
        questions_created = 0
        questions_updated = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            year_val = row[col['학년도']]
            exam_type_raw = row[col['시험종류']]
            subject_name = row[col['과목명']]
            grade_val = row[col['학년']]
            number_val = row[col['문제번호']]
            text_val = row[col['문제']]
            choice_1 = row[col['1항']]
            choice_2 = row[col['2항']]
            choice_3 = row[col['3항']]
            choice_4 = row[col['4항']]
            answer_val = row[col['답안']]

            if not year_val or not number_val or not text_val:
                continue

            year = int(year_val)
            number = int(number_val)
            grade = int(grade_val) if grade_val else 3

            exam_type = EXAM_TYPE_MAP.get(str(exam_type_raw).strip(), str(exam_type_raw).strip())
            if exam_type not in ('중간', '기말', '계절'):
                self.stderr.write(self.style.WARNING(
                    f'  알 수 없는 시험종류: {exam_type_raw} (건너뜀)'
                ))
                continue

            # Exam
            exam, created = Exam.objects.get_or_create(
                year=year,
                exam_type=exam_type,
            )
            if created:
                exams_created += 1

            # Subject
            subject, _ = Subject.objects.get_or_create(
                name=subject_name,
                grade=grade,
                defaults={'semester': 1, 'department': '농학과', 'category': '전공'},
            )

            # Answer: 문자열로 변환 (1~4 단일답, A~K 중복답, 0 미확인)
            answer_str = str(answer_val).strip() if answer_val else '0'
            if answer_str in ('1', '2', '3', '4'):
                answer = answer_str
            elif answer_str.upper() in MULTI_ANSWER_MAP:
                answer = MULTI_ANSWER_MAP[answer_str.upper()]
            elif ',' in answer_str:
                # 이미 "1,2" 형식인 경우 그대로 사용
                answer = answer_str
            else:
                answer = '0'

            # Question (update_or_create: 화학식 변환 반영)
            _, created = Question.objects.update_or_create(
                subject=subject,
                year=year,
                number=number,
                defaults={
                    'text': convert_formulas(text_val),
                    'choice_1': convert_formulas(choice_1),
                    'choice_2': convert_formulas(choice_2),
                    'choice_3': convert_formulas(choice_3),
                    'choice_4': convert_formulas(choice_4),
                    'answer': answer,
                },
            )
            if created:
                questions_created += 1
            else:
                questions_updated += 1

        wb.close()
        return exams_created, questions_created, questions_updated
