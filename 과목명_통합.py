"""
과목명 중복 통합 및 엑셀 재생성 스크립트
- 같은 과목인데 이름이 다른 것들을 통합
- JSON 업데이트 & 엑셀 재생성
"""

import json
import re

INPUT_JSON = "카페글_시험문제.json"
OUTPUT_JSON = "카페글_시험문제.json"  # 덮어쓰기
OUTPUT_EXCEL = "카페글_시험문제.xlsx"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 과목명 통합 매핑 (잘못된 이름 → 정식 이름)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SUBJECT_MAPPING = {
    # 해충방제학
    "해충 방제학": "해충방제학",

    # 숲과삶
    "숲과 삶": "숲과삶",

    # 잡초방제학
    "잡초방제": "잡초방제학",

    # 식용작물학1
    "식작1": "식용작물학1",
    "식용작물": "식용작물학1",  # 문맥상 식용작물학1인 경우 많음
    "식용작물학(벼)": "식용작물학1",

    # 식용작물학2
    "식작2": "식용작물학2",
    "식용작물2": "식용작물학2",

    # 생물통계학
    "생물통계": "생물통계학",
    "통계학": "생물통계학",

    # 반려동물학
    "반려동물": "반려동물학",

    # 동물사료학 / 사료학
    "사료학": "동물사료학",

    # 시설원예학
    "시설원예": "시설원예학",

    # 재배학원론
    "재배학 원론": "재배학원론",
    "재배학": "재배학원론",

    # 농업경영학
    "경영학": "농업경영학",

    # 환경친화형농업
    # (이미 정상)
}


def normalize_subject(name):
    """과목명 정규화"""
    name = name.strip()
    # 매핑 테이블에 있으면 변환
    if name in SUBJECT_MAPPING:
        return SUBJECT_MAPPING[name]
    return name


def save_to_excel(all_questions, filepath):
    """추출된 문제를 엑셀로 저장"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "시험문제"

    # 헤더
    headers = ["과목", "일자", "문제번호", "문제", "보기", "답", "출처(글번호)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

    # 데이터 입력
    row = 2
    for q in all_questions:
        ws.cell(row=row, column=1, value=q.get("과목", ""))
        ws.cell(row=row, column=2, value=q.get("일자", ""))
        ws.cell(row=row, column=3, value=q.get("문제번호", ""))
        ws.cell(row=row, column=4, value=q.get("문제", ""))
        ws.cell(row=row, column=5, value=q.get("보기", ""))
        ws.cell(row=row, column=6, value=q.get("답", ""))
        ws.cell(row=row, column=7, value=q.get("출처", ""))
        row += 1

    # 열 너비 조정
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 12

    wb.save(filepath)


def main():
    print("=" * 60)
    print("📝 과목명 통합 스크립트")
    print("=" * 60)

    # 1. JSON 읽기
    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)

    questions = data["questions"]
    print(f"\n📂 총 {len(questions)}개 문제 로드")

    # 2. 과목명 통합 전 통계
    before_subjects = {}
    for q in questions:
        subj = q.get("과목", "미분류")
        before_subjects[subj] = before_subjects.get(subj, 0) + 1

    print(f"   통합 전 과목 수: {len(before_subjects)}개")

    # 3. 과목명 통합
    changed_count = 0
    change_log = {}
    for q in questions:
        old_name = q.get("과목", "미분류")
        new_name = normalize_subject(old_name)
        if old_name != new_name:
            q["과목"] = new_name
            changed_count += 1
            key = f"{old_name} → {new_name}"
            change_log[key] = change_log.get(key, 0) + 1

    # 4. 통합 후 통계
    after_subjects = {}
    for q in questions:
        subj = q.get("과목", "미분류")
        after_subjects[subj] = after_subjects.get(subj, 0) + 1

    print(f"   통합 후 과목 수: {len(after_subjects)}개")
    print(f"   변경된 문제 수: {changed_count}개")

    if change_log:
        print(f"\n{'─' * 50}")
        print("변경 내역:")
        print(f"{'─' * 50}")
        for change, count in sorted(change_log.items(), key=lambda x: -x[1]):
            print(f"  {change}: {count}개")

    # 5. 최종 과목 목록 출력
    print(f"\n{'=' * 50}")
    print(f"최종 과목별 문제 수 ({len(after_subjects)}개 과목):")
    print(f"{'=' * 50}")
    for k, v in sorted(after_subjects.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}개")

    # 6. JSON 저장
    data["questions"] = questions
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\n💾 JSON 저장 완료: {OUTPUT_JSON}")

    # 7. 엑셀 재생성
    save_to_excel(questions, OUTPUT_EXCEL)
    print(f"💾 엑셀 저장 완료: {OUTPUT_EXCEL}")

    print(f"\n{'=' * 60}")
    print(f"✅ 완료! {len(questions)}개 문제, {len(after_subjects)}개 과목")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
