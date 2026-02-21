"""
카페글_결과.txt에서 시험 문제를 추출하여 엑셀로 변환하는 스크립트
- Gemini API를 사용하여 비정형 텍스트에서 문제 추출
- 과목, 일자, 문제번호, 문제내용, 보기, 답 형태로 구조화
"""

import os
import re
import json
import time
from dotenv import load_dotenv

load_dotenv()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 설정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
INPUT_FILE = "카페글_결과.txt"
OUTPUT_EXCEL = "카페글_시험문제.xlsx"
OUTPUT_JSON = "카페글_시험문제.json"  # 중간 저장용
GEMINI_MODEL = "gemini-2.5-flash"
API_DELAY = 2  # API 호출 간격 (초)


def split_articles(filepath):
    """
    카페글_결과.txt를 개별 글로 분리
    """
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    # 글 단위로 분리 (━━━ 구분선 기준)
    pattern = r"━{10,}\s*\n📌 \[(\d+)번째 글\]\s*\n━{10,}\s*\n"
    parts = re.split(pattern, content)

    articles = []
    # parts[0]은 첫 구분선 이전 (빈 문자열)
    # parts[1] = 글번호, parts[2] = 글 내용, parts[3] = 글번호, parts[4] = 글 내용, ...
    for i in range(1, len(parts) - 1, 2):
        article_num = int(parts[i])
        article_content = parts[i + 1].strip()

        # 제목, 작성자, 날짜 추출
        lines = article_content.split("\n")
        title = ""
        author = ""
        date = ""
        body = ""

        for j, line in enumerate(lines):
            line = line.strip()
            if line.startswith("제목:"):
                title = line[3:].strip()
            elif line.startswith("작성자:"):
                author = line[4:].strip()
            elif line.startswith("작성일:"):
                date = line[4:].strip()
            elif line.startswith("=" * 20):
                # 본문 시작
                body_lines = []
                for k in range(j + 1, len(lines)):
                    body_lines.append(lines[k])
                body = "\n".join(body_lines)
                break

        articles.append({
            "num": article_num,
            "title": title,
            "author": author,
            "date": date,
            "body": body,
        })

    return articles


def extract_questions_with_gemini(article):
    """
    Gemini API를 사용하여 게시글에서 시험 문제를 추출
    """
    from google import genai

    client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

    prompt = f"""다음은 한국방송통신대학교 농학과 학생이 기말시험 후기로 올린 카페 게시글입니다.
이 글에서 시험에 출제된 문제를 추출해주세요.

**추출 규칙:**
1. 실제 시험 문제만 추출합니다. 일반 소감, 공부 방법, 인사말은 무시합니다.
2. 문제가 없는 글(단순 소감, 질문글 등)은 빈 배열 []을 반환합니다.
3. 과목명은 제목이나 본문에서 추출합니다. 하나의 글에 여러 과목이 있을 수 있습니다.
4. 댓글에 추가 문제 정보가 있으면 함께 추출합니다.
5. 답이 명시되어 있으면 추출하고, 없으면 빈 문자열로 남깁니다.
6. 보기(선택지)가 있으면 ①②③④ 형태로 정리합니다. 없으면 빈 문자열.
7. 시험 일자는 작성일 기준으로 추정합니다 (보통 시험 당일이나 다음날 작성).

**출력 형식 (JSON 배열):**
```json
[
  {{
    "과목": "과목명",
    "일자": "YYYY.MM.DD",
    "문제번호": 1,
    "문제": "문제 내용",
    "보기": "①보기1 ②보기2 ③보기3 ④보기4",
    "답": "정답 내용"
  }}
]
```

문제가 없으면 빈 배열 `[]`을 반환하세요.
반드시 JSON 형식만 출력하세요. 설명이나 마크다운 코드블록 없이 순수 JSON만.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
제목: {article['title']}
작성일: {article['date']}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{article['body'][:4000]}
"""

    try:
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
        )
        text = response.text.strip()

        # JSON 파싱 (마크다운 코드블록 제거)
        text = re.sub(r"^```json\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        text = text.strip()

        questions = json.loads(text)
        return questions

    except json.JSONDecodeError as e:
        print(f"  ⚠️  JSON 파싱 실패 (글 #{article['num']}): {e}")
        print(f"  응답: {text[:200]}...")
        return []
    except Exception as e:
        print(f"  ❌ API 오류 (글 #{article['num']}): {e}")
        return []


def save_to_excel(all_questions, filepath):
    """
    추출된 문제를 엑셀로 저장
    """
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 설치 필요: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "시험문제"

    # 헤더
    headers = ["과목", "일자", "문제번호", "문제", "보기", "답", "출처(글번호)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    # 데이터 입력
    row = 2
    for q in all_questions:
        ws.cell(row=row, column=1, value=q.get("과목", ""))
        ws.cell(row=row, column=2, value=q.get("일자", ""))
        ws.cell(row=row, column=3, value=q.get("문제번호", ""))
        ws.cell(row=row, column=4, value=q.get("문제", ""))
        ws.cell(row=row, column=5, value=q.get("보기", ""))
        ws.cell(row=row, column=6, value=q.get("답", ""))
        ws.cell(row=row, column=7, value=q.get("출처", ""))
        row += 1

    # 열 너비 조정
    ws.column_dimensions["A"].width = 20  # 과목
    ws.column_dimensions["B"].width = 14  # 일자
    ws.column_dimensions["C"].width = 10  # 문제번호
    ws.column_dimensions["D"].width = 60  # 문제
    ws.column_dimensions["E"].width = 60  # 보기
    ws.column_dimensions["F"].width = 30  # 답
    ws.column_dimensions["G"].width = 12  # 출처

    wb.save(filepath)
    print(f"\n✅ 엑셀 저장 완료: {filepath}")
    print(f"   총 {row - 2}개 문제")


def main():
    print("=" * 60)
    print("📝 카페글 시험문제 추출기 (Gemini LLM)")
    print("=" * 60)

    # 1. 글 분리
    print(f"\n📂 {INPUT_FILE} 읽는 중...")
    articles = split_articles(INPUT_FILE)
    print(f"   총 {len(articles)}개 글 발견")

    # 2. 기존 진행 상황 확인 (이어서 처리 가능)
    all_questions = []
    processed_nums = set()

    if os.path.exists(OUTPUT_JSON):
        with open(OUTPUT_JSON, "r", encoding="utf-8") as f:
            saved_data = json.load(f)
            all_questions = saved_data.get("questions", [])
            processed_nums = set(saved_data.get("processed", []))
        print(f"   기존 진행: {len(processed_nums)}개 글 처리됨, {len(all_questions)}개 문제 추출됨")

        choice = input("   이어서 처리할까요? (1: 이어서 / 2: 처음부터): ").strip()
        if choice == "2":
            all_questions = []
            processed_nums = set()

    # 3. Gemini로 문제 추출
    print(f"\n🤖 Gemini API로 문제 추출 시작...")
    print(f"   모델: {GEMINI_MODEL}")
    print("-" * 60)

    for article in articles:
        if article["num"] in processed_nums:
            continue

        print(f"\n📌 [{article['num']}번째 글] {article['title']}")

        # 본문이 너무 짧으면 (50자 미만) 문제가 없을 가능성 높음
        if len(article["body"].strip()) < 50:
            print(f"   ⏭️  본문 너무 짧음 - 건너뜀")
            processed_nums.add(article["num"])
            continue

        questions = extract_questions_with_gemini(article)

        if questions:
            # 출처 정보 추가
            for q in questions:
                q["출처"] = f"글#{article['num']}"
                # 일자가 비어있으면 작성일 사용
                if not q.get("일자"):
                    # 작성일에서 날짜만 추출 (예: "2025.12.15. 18:18" → "2025.12.15")
                    date_match = re.match(r"(\d{4}\.\d{2}\.\d{2})", article["date"])
                    if date_match:
                        q["일자"] = date_match.group(1)

            all_questions.extend(questions)
            print(f"   ✅ {len(questions)}개 문제 추출")
        else:
            print(f"   ⏭️  추출된 문제 없음")

        processed_nums.add(article["num"])

        # 중간 저장 (매 5개 글마다)
        if len(processed_nums) % 5 == 0:
            with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
                json.dump(
                    {"questions": all_questions, "processed": list(processed_nums)},
                    f,
                    ensure_ascii=False,
                    indent=2,
                )
            print(f"   💾 중간 저장 ({len(all_questions)}개 문제)")

        # API 속도 제한
        time.sleep(API_DELAY)

    # 4. 최종 JSON 저장
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(
            {"questions": all_questions, "processed": list(processed_nums)},
            f,
            ensure_ascii=False,
            indent=2,
        )
    print(f"\n💾 JSON 저장 완료: {OUTPUT_JSON}")

    # 5. 엑셀 변환
    if all_questions:
        save_to_excel(all_questions, OUTPUT_EXCEL)
    else:
        print("\n⚠️  추출된 문제가 없습니다.")

    print("\n" + "=" * 60)
    print("✅ 완료!")
    print(f"   처리된 글: {len(processed_nums)}개")
    print(f"   추출된 문제: {len(all_questions)}개")
    print("=" * 60)


if __name__ == "__main__":
    main()
